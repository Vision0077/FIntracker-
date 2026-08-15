"""
Day 17: Excel Bank Statement Parser (.xlsx / .xls)
=====================================================
Parses Excel bank statement exports using openpyxl.

Strategy:
  1. Open workbook in read-only mode (memory efficient for large files).
  2. Find the worksheet most likely to contain transactions:
     - Prefer sheets named 'Statement', 'Transactions', 'Sheet1', etc.
     - Fall back to the first sheet.
  3. Scan rows to locate the header row:
     - Same alias-based approach as csv_parser and pdf_parser.
     - Skip leading metadata rows (bank name, account number, date range blurb).
  4. Parse data rows using the shared helper functions from csv_parser.

Reuses ALL helpers from csv_parser — zero code duplication:
  _parse_date, _parse_amount, _detect_payment_method, _build_transaction,
  _find_col, _normalise_header, and all column alias lists.

Usage:
    txns, errors = parse_excel(file_bytes, filename="hdfc_jun.xlsx", user_id=uid)
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from app.services.csv_parser import (
    DATE_COLS,
    DESC_COLS,
    DEBIT_COLS,
    CREDIT_COLS,
    AMOUNT_COLS,
    TYPE_COLS,
    PAYMENT_METHOD_COLS,
    CATEGORY_COLS,
    _normalise_header,
    _find_col,
    _parse_date,
    _parse_amount,
    _detect_payment_method,
    _build_transaction,
    ParseError,
)
from app.schemas.transaction import TransactionCreate


# ---------------------------------------------------------------------------
# Preferred sheet name keywords (priority order)
# ---------------------------------------------------------------------------

PREFERRED_SHEET_NAMES = [
    "statement", "transactions", "transaction", "txn",
    "account statement", "bank statement", "sheet1",
    "data", "details",
]


def _pick_sheet(workbook):
    """
    Choose the best sheet to parse from a workbook.
    Priority: name matches PREFERRED_SHEET_NAMES (case-insensitive) → first sheet.
    """
    sheets = workbook.sheetnames
    if not sheets:
        return None
    sheets_lower = [s.lower() for s in sheets]
    for keyword in PREFERRED_SHEET_NAMES:
        for i, name in enumerate(sheets_lower):
            if keyword in name:
                return workbook[sheets[i]]
    return workbook[sheets[0]]


def _cell_to_str(cell_value) -> str:
    """
    Convert an openpyxl cell value to a clean string.
    Handles: str, int, float, Decimal, datetime, date, None.
    """
    if cell_value is None:
        return ""
    if isinstance(cell_value, (datetime, date)):
        # Return ISO date string — _parse_date handles YYYY-MM-DD
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, float):
        # Avoid '1250.0' when the value is a whole number
        if cell_value == int(cell_value):
            return str(int(cell_value))
        return str(round(cell_value, 2))
    return str(cell_value).strip()


# ---------------------------------------------------------------------------
# Header scanner
# ---------------------------------------------------------------------------


def _find_header_row(sheet, max_scan: int = 20) -> Optional[int]:
    """
    Scan the first `max_scan` rows to find the one most likely to be the header.

    Scoring: count how many cells in the row contain any column alias keyword.
    Row with score >= 2 and the highest score wins.
    Returns the 1-based row index, or None if no header found.
    """
    all_aliases = DATE_COLS + DESC_COLS + DEBIT_COLS + CREDIT_COLS + AMOUNT_COLS

    best_row = None
    best_score = 0

    for row_idx in range(1, min(max_scan + 1, sheet.max_row + 1)):
        row_cells = [_cell_to_str(sheet.cell(row_idx, col).value) for col in range(1, sheet.max_column + 1)]
        norm = [_normalise_header(c) for c in row_cells]
        score = sum(1 for cell in norm if any(alias in cell for alias in all_aliases))
        if score > best_score:
            best_score = score
            best_row = row_idx

    if best_score >= 2:
        return best_row
    # Fallback: first non-empty row
    for row_idx in range(1, min(max_scan + 1, sheet.max_row + 1)):
        row_cells = [_cell_to_str(sheet.cell(row_idx, col).value) for col in range(1, sheet.max_column + 1)]
        if any(c for c in row_cells):
            return row_idx
    return None


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def parse_excel(
    file_bytes: bytes,
    filename: str,
    user_id: uuid.UUID,
) -> tuple[list[TransactionCreate], list[str]]:
    """
    Parse an Excel (.xlsx/.xls) bank statement.

    Returns:
        (transactions, errors)
        transactions — list[TransactionCreate] ready to insert
        errors       — list[str] per-row error messages for bad rows
    """
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl is not installed. Run: pip install openpyxl"]

    transactions: list[TransactionCreate] = []
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Could not open Excel file: {e}"]

    sheet = _pick_sheet(wb)
    if sheet is None:
        return [], ["Excel file has no sheets."]

    # Find header row
    header_row_idx = _find_header_row(sheet)
    if header_row_idx is None:
        return [], [f"Could not identify header row in sheet '{sheet.title}'."]

    # Extract headers
    headers = [
        _cell_to_str(sheet.cell(header_row_idx, col).value)
        for col in range(1, sheet.max_column + 1)
    ]

    # Detect column indices
    date_idx   = _find_col(headers, DATE_COLS)
    desc_idx   = _find_col(headers, DESC_COLS)
    debit_idx  = _find_col(headers, DEBIT_COLS)
    credit_idx = _find_col(headers, CREDIT_COLS)
    amount_idx = _find_col(headers, AMOUNT_COLS)
    type_idx   = _find_col(headers, TYPE_COLS)
    method_idx = _find_col(headers, PAYMENT_METHOD_COLS)
    cat_idx    = _find_col(headers, CATEGORY_COLS)

    if date_idx is None or desc_idx is None:
        errors.append(
            f"No date/description columns found. Headers: {headers}"
        )
        wb.close()
        return transactions, errors

    has_debit_credit = debit_idx is not None and credit_idx is not None
    has_amount = amount_idx is not None

    if not has_debit_credit and not has_amount:
        errors.append(f"No amount column found. Headers: {headers}")
        wb.close()
        return transactions, errors

    # Parse data rows
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        # Read all cells in row
        row = [
            _cell_to_str(sheet.cell(row_idx, col).value)
            for col in range(1, sheet.max_column + 1)
        ]

        # Skip blank rows
        if not any(c for c in row):
            continue

        # Pad if necessary
        max_col = max(
            date_idx, desc_idx,
            debit_idx or 0, credit_idx or 0,
            amount_idx or 0,
        )
        while len(row) <= max_col:
            row.append("")

        # --- Date ---
        raw_date = row[date_idx].strip()
        if not raw_date:
            continue
        txn_date = _parse_date(raw_date)
        if txn_date is None:
            continue  # summary/total row

        # --- Description ---
        description = row[desc_idx].strip() or "Unknown transaction"

        # --- Amount and type ---
        if has_debit_credit:
            debit  = _parse_amount(row[debit_idx])  if debit_idx  is not None else None
            credit = _parse_amount(row[credit_idx]) if credit_idx is not None else None

            if debit and debit > 0:
                amount, txn_type = debit, "EXPENSE"
            elif credit and credit > 0:
                amount, txn_type = credit, "INCOME"
            else:
                continue
        else:
            raw_amt = row[amount_idx].strip() if amount_idx is not None else ""
            parsed = _parse_amount(raw_amt)
            if parsed is None or parsed == 0:
                continue
            abs_amount = abs(parsed)
            if type_idx is not None and row[type_idx].strip():
                raw_type = row[type_idx].strip().upper()
                txn_type = "INCOME" if raw_type in ("INCOME", "CR", "CREDIT") else "EXPENSE"
            else:
                txn_type = "EXPENSE" if parsed < 0 else "INCOME"
            amount = abs_amount

        # --- Payment method ---
        if method_idx is not None and len(row) > method_idx and row[method_idx].strip():
            raw_method = row[method_idx].strip().upper()
            valid_methods = {"UPI", "CARD", "WALLET", "SUBSCRIPTION", "CASH"}
            payment_method = raw_method if raw_method in valid_methods else _detect_payment_method(description)
        else:
            payment_method = _detect_payment_method(description)

        # --- Category ---
        category = None
        if cat_idx is not None and len(row) > cat_idx and row[cat_idx].strip():
            category = row[cat_idx].strip().upper()

        # --- Build ---
        try:
            txn = _build_transaction(
                row_num=row_idx,
                txn_date=txn_date,
                description=description,
                amount=amount,
                txn_type=txn_type,
                payment_method=payment_method,
                category=category,
                user_id=user_id,
            )
            transactions.append(txn)
        except ParseError as exc:
            errors.append(str(exc))
        except Exception as exc:
            errors.append(f"Row {row_idx}: unexpected error — {exc}")

    wb.close()
    return transactions, errors
